# ============================================================
# ✅ FlashProd KPI Suite (STREAMLIT) — VERSION COMPLÈTE À JOUR
# ✅ Ajouts TL:
#   - Agents présents (>=10 min) en premier
#   - Agents < 13 juste après Productivité
#   - Heures Meeting / Training / OJT / Coaching total
#   - % Coaching vs Connecté
# ✅ Email: bandeau + table TL avec % coaching + counts
# ✅ Output: Excel final + Email HTML + FLOP optionnel
# ============================================================

import streamlit as st
import pandas as pd
import numpy as np
import re
from io import BytesIO
from datetime import datetime, time, timedelta

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


# =========================
# Streamlit config
# =========================
st.set_page_config(page_title="FlashProd KPI Suite", layout="wide")


# =========================
# Helpers
# =========================
def read_excel_any(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    bio = BytesIO(uploaded_file.getvalue())
    if name.endswith(".xls"):
        return pd.read_excel(bio, engine="xlrd")
    return pd.read_excel(bio, engine="openpyxl")


def normalize_log_series(s: pd.Series) -> pd.Series:
    s = s.astype(str).str.strip()
    s = s.str.replace(r"\.0$", "", regex=True)
    s = s.str.replace(r"[^\d]", "", regex=True)
    s = s.replace({"": np.nan, "nan": np.nan, "None": np.nan})
    return s


def extract_log_any(text) -> str | None:
    if text is None or (isinstance(text, float) and np.isnan(text)):
        return None
    t = str(text)
    m = re.search(r"(?:agent|log|id)\D*(\d{4,6})", t, flags=re.IGNORECASE)
    if m:
        return m.group(1)
    m = re.search(r"(\d{4,6})", t)
    if m:
        return m.group(1)
    return None


def to_td(val):
    if val is None:
        return timedelta(0)
    if isinstance(val, str):
        v = val.strip()
        if v in ("-", "", "0", "00:00:00"):
            return timedelta(0)
        mm = re.match(r"(\d+)h(\d+)'(\d+)", v)
        if mm:
            h, m, s = mm.groups()
            return timedelta(hours=int(h), minutes=int(m), seconds=int(s))
    try:
        td = pd.to_timedelta(val, errors="coerce")
        if pd.isna(td):
            return timedelta(0)
        return td
    except Exception:
        return timedelta(0)


def fmt_hms(td: timedelta) -> str:
    total = int(td.total_seconds())
    sign = "-" if total < 0 else ""
    total = abs(total)
    h = total // 3600
    m = (total % 3600) // 60
    s = total % 60
    return f"{sign}{h:02d}:{m:02d}:{s:02d}"


def safe_int(x):
    try:
        if x is None or (isinstance(x, float) and np.isnan(x)):
            return 0
        s = str(x).strip()
        if s in ("-", ""):
            return 0
        return int(float(s))
    except Exception:
        return 0


def to_seconds(val):
    if val in ("-", None, ""):
        return None
    if isinstance(val, time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, timedelta):
        return int(val.total_seconds())
    if isinstance(val, (int, float)):
        return int(val)
    if isinstance(val, str) and ":" in val:
        parts = val.split(":")
        if len(parts) >= 2:
            h = int(parts[0])
            m = int(parts[1])
            s = int(parts[2]) if len(parts) >= 3 else 0
            return h * 3600 + m * 60 + s
    return None


def time_to_seconds(val):
    if pd.isna(val):
        return None
    if isinstance(val, timedelta):
        return val.total_seconds()
    if isinstance(val, time):
        return val.hour * 3600 + val.minute * 60 + val.second
    if isinstance(val, str) and ":" in val:
        parts = val.split(":")
        if len(parts) >= 2:
            h = int(parts[0])
            m = int(parts[1])
            s = int(parts[2]) if len(parts) >= 3 else 0
            return h * 3600 + m * 60 + s
    return None


def sec_to_hms(sec):
    if sec is None or (isinstance(sec, float) and np.isnan(sec)):
        return "-"
    return str(timedelta(seconds=int(float(sec))))


def parse_percent_any(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    if isinstance(v, (int, float)):
        f = float(v)
        return f if f <= 1.5 else f / 100.0
    s = str(v).strip().replace(",", ".")
    if s in ("-", "", "nan", "None"):
        return np.nan
    if s.endswith("%"):
        try:
            return float(s[:-1]) / 100.0
        except Exception:
            return np.nan
    try:
        f = float(s)
        return f if f <= 1.5 else f / 100.0
    except Exception:
        return np.nan


def parse_hms_any_to_seconds(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    if isinstance(v, timedelta):
        return float(v.total_seconds())
    if isinstance(v, time):
        return float(v.hour * 3600 + v.minute * 60 + v.second)
    s = str(v).strip()
    if s in ("-", "", "nan", "None"):
        return np.nan
    td = pd.to_timedelta(s, errors="coerce")
    if pd.isna(td):
        return np.nan
    return float(td.total_seconds())


def kpi_or_dash_num(v, digits=1):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    try:
        return str(round(float(v), digits)).replace(".", ",")
    except Exception:
        return "-"


def kpi_or_dash_pct(v, digits=1):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "-"
    try:
        return f"{round(float(v)*100, digits)}%".replace(".", ",")
    except Exception:
        return "-"


# =========================
# CODE 1: Nettoyage BRUT
# =========================
def code1_clean_raw(df: pd.DataFrame) -> pd.DataFrame:
    df = df.dropna(how="all").dropna(axis=1, how="all")

    rename_map = {}
    if "Unnamed: 0" in df.columns: rename_map["Unnamed: 0"] = "Nom Agent"
    if "Unnamed: 1" in df.columns: rename_map["Unnamed: 1"] = "Etat"
    if "Unnamed: 4" in df.columns: rename_map["Unnamed: 4"] = "Occurances"
    if "Unnamed: 6" in df.columns: rename_map["Unnamed: 6"] = "Temps total"
    df = df.rename(columns=rename_map)

    for col in ["Nom Agent", "Etat", "Occurances", "Temps total"]:
        if col not in df.columns:
            df[col] = np.nan

    df = df[df["Etat"].notna()]

    df.insert(0, "Log Téléphonie1", df["Nom Agent"].apply(extract_log_any))
    df["Log Téléphonie1"] = normalize_log_series(df["Log Téléphonie1"])

    df["Etat"] = df["Etat"].astype(str).str.strip()
    df["Etat"] = df["Etat"].replace({
        "Attente": "Attente global",
        "Pause": "Pause global",
        "Preview": "Histo Mailing"
    })
    df = df[df["Etat"].astype(str).str.lower() != "en attente"]

    df["Temps total"] = df["Temps total"].apply(
        lambda x: fmt_hms(to_td(x)) if str(x).strip() not in ("-", "", "nan") else "-"
    )

    def rename_second_pause(group):
        pauses = group[group["Etat"] == "Pause"].index
        if len(pauses) > 1:
            group.loc[pauses[1:], "Etat"] = "Pause générique"
        return group

    df = df.groupby("Log Téléphonie1", group_keys=False).apply(rename_second_pause)
    return df


# =========================
# CODE 2: KPI Agent / Pivot
# =========================
def code2_build_agent_kpis(df_clean: pd.DataFrame) -> pd.DataFrame:
    df2 = df_clean.copy()

    if "Log Téléphonie1" not in df2.columns:
        df2["Log Téléphonie1"] = df2["Nom Agent"].apply(extract_log_any)
    df2["Log Téléphonie1"] = normalize_log_series(df2["Log Téléphonie1"])
    df2 = df2[df2["Log Téléphonie1"].notna()].copy()

    unique_etats = df2["Etat"].dropna().unique()
    agent_base = df2[["Log Téléphonie1"]].drop_duplicates().reset_index(drop=True)

    for etat in unique_etats:
        agent_base[f"{etat} - Occurence"] = "-"
        agent_base[f"{etat} - Temps total"] = "-"

    for i, row in agent_base.iterrows():
        log = str(row["Log Téléphonie1"]).strip()
        sub_df = df2[df2["Log Téléphonie1"].astype(str).str.strip() == log]
        if sub_df.empty:
            continue

        for etat in unique_etats:
            bloc = sub_df[sub_df["Etat"] == etat]
            if bloc.empty:
                continue

            occ_col = "Occurances" if "Occurances" in bloc.columns else ("Occurrences" if "Occurrences" in bloc.columns else None)
            occ_sum = safe_int(bloc[occ_col].fillna(0).sum()) if occ_col else 0

            tm_sum = timedelta(0)
            for v in bloc["Temps total"].tolist():
                tm_sum += to_td(v)

            agent_base.at[i, f"{etat} - Occurence"] = occ_sum if occ_sum != 0 else "-"
            agent_base.at[i, f"{etat} - Temps total"] = fmt_hms(tm_sum) if tm_sum != timedelta(0) else "-"

    etat_presence = ["Attente global", "Traitement", "Post-travail", "Pause global"]

    def total_presence_row(r):
        total = timedelta(0)
        for e in etat_presence:
            col = f"{e} - Temps total"
            if col in r:
                total += to_td(r[col])
        return fmt_hms(total)

    agent_base["Temps Total présence"] = agent_base.apply(total_presence_row, axis=1)

    def total_travail_row(r):
        return fmt_hms(to_td(r.get("Traitement - Temps total", "-")) + to_td(r.get("Post-travail - Temps total", "-")))

    agent_base.insert(
        agent_base.columns.get_loc("Temps Total présence") + 1,
        "Temps total Travail",
        agent_base.apply(total_travail_row, axis=1)
    )

    def taux_occupation_row(r):
        travail = to_td(r["Temps total Travail"]).total_seconds()
        presence = to_td(r["Temps Total présence"]).total_seconds()
        if presence <= 0:
            return "0.00%"
        return f"{(travail / presence) * 100:.2f}%"

    agent_base.insert(
        agent_base.columns.get_loc("Temps total Travail") + 1,
        "Taux d'occupation",
        agent_base.apply(taux_occupation_row, axis=1)
    )

    def productivite_row(r):
        occ = safe_int(r.get("Traitement - Occurence", 0))
        presence_h = to_td(r["Temps Total présence"]).total_seconds() / 3600
        if presence_h <= 0:
            return np.nan
        return round(occ / presence_h, 2)

    agent_base["Productivité"] = agent_base.apply(productivite_row, axis=1)

    def calc_dmc(r):
        occ = safe_int(r.get("Traitement - Occurence", 0))
        if occ <= 0:
            return "-"
        total = to_td(r.get("Traitement - Temps total", "00:00:00")) / occ
        return fmt_hms(total)

    agent_base["DMC"] = agent_base.apply(calc_dmc, axis=1)

    def calc_dmt(r):
        occ = safe_int(r.get("Traitement - Occurence", 0))
        if occ <= 0:
            return "-"
        total = (to_td(r.get("Traitement - Temps total", "00:00:00")) +
                 to_td(r.get("Post-travail - Temps total", "00:00:00"))) / occ
        return fmt_hms(total)

    agent_base["DMT"] = agent_base.apply(calc_dmt, axis=1)

    return agent_base


# =========================
# CODE 3: Merge COMPO + Rapport
# =========================
def code3_merge_compo(df_compo: pd.DataFrame, df_rapport: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    if "Log Téléphonie1" not in df_compo.columns:
        raise ValueError("COMPO: colonne 'Log Téléphonie1' introuvable.")
    if "Log Téléphonie1" not in df_rapport.columns:
        raise ValueError("RAPPORT: colonne 'Log Téléphonie1' introuvable.")

    df_compo = df_compo.copy()
    df_rapport = df_rapport.copy()

    df_compo["Log Téléphonie1"] = normalize_log_series(df_compo["Log Téléphonie1"])
    df_rapport["Log Téléphonie1"] = normalize_log_series(df_rapport["Log Téléphonie1"])

    before_compo = len(df_compo)
    df_compo = df_compo[df_compo["Log Téléphonie1"].notna()].copy()
    df_rapport = df_rapport[df_rapport["Log Téléphonie1"].notna()].copy()

    df_merged = pd.merge(df_compo, df_rapport, on="Log Téléphonie1", how="left", indicator=True)

    stats = {
        "compo_rows": before_compo,
        "compo_rows_non_null_log": len(df_compo),
        "rapport_rows": len(df_rapport),
        "matched_rows": int((df_merged["_merge"] == "both").sum()),
        "unmatched_rows": int((df_merged["_merge"] != "both").sum()),
    }
    df_merged.drop(columns=["_merge"], inplace=True)

    if "Nom Agent" in df_merged.columns:
        df_merged = df_merged[df_merged["Nom Agent"].notna() & (df_merged["Nom Agent"] != "")]

    colonnes_apres_ops = ["Temps Total présence", "Temps total Travail", "Taux d'occupation", "Productivité", "DMC", "DMT"]
    ops_index = df_merged.columns.get_loc("OPS") + 1 if "OPS" in df_merged.columns else 0
    colonnes_sans = [c for c in df_merged.columns if c not in colonnes_apres_ops]
    nouvel_ordre = colonnes_sans[:ops_index] + [c for c in colonnes_apres_ops if c in df_merged.columns] + colonnes_sans[ops_index:]

    return df_merged[nouvel_ordre], stats


# =========================
# CODE 4: Flash Prod Agent (Excel)
# =========================
def code4_build_flash_agent_excel(df_final3: pd.DataFrame) -> bytes:
    df4 = df_final3.copy()
    df4 = df4.fillna("-")

    if "Appel entrant - Occurence" in df4.columns:
        df4 = df4.rename(columns={"Appel entrant - Occurence": "Appels entrants"})

    columns_to_keep = [
        "Matricule RH","Log Téléphonie1","Nom Agent","File","Tls","OPS",
        "Temps Total présence","Temps total Travail","Taux d'occupation",
        "Productivité","DMC","DMT","Attente global - Temps total",
        "Appels entrants","Appel entrant - Temps total",
        "Post-travail - Temps total","Break - Temps total",
        "BUG IT - Temps total","Meeting - Temps total",
        "Pause générique - Temps total","Training - Temps total",
        "Back Office - Temps total","Détachement - Temps total",
        "Call Back - Temps total","Mailing - Temps total","OJT - Temps total"
    ]
    df4 = df4[[c for c in columns_to_keep if c in df4.columns]]

    buf = BytesIO()
    df4.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb.active
    ws.title = "Flash Prod Agent"

    headers = {cell.value: cell.column for cell in ws[1]}

    text_cols = ["Matricule RH","Log Téléphonie1","Nom Agent","File","Tls","OPS"]
    time_cols = [
        "Temps Total présence","Temps total Travail","DMC","DMT",
        "Attente global - Temps total","Appel entrant - Temps total",
        "Post-travail - Temps total","Break - Temps total",
        "BUG IT - Temps total","Meeting - Temps total",
        "Pause générique - Temps total","Training - Temps total",
        "Back Office - Temps total","Détachement - Temps total",
        "Call Back - Temps total","Mailing - Temps total","OJT - Temps total"
    ]

    for name in text_cols:
        if name in headers:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, headers[name]).number_format = "@"

    for name in time_cols:
        if name in headers:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, headers[name]).number_format = "hh:mm:ss"

    # add Taux Post-travail
    if "Post-travail - Temps total" in headers and "Temps total Travail" in headers:
        col_post = headers["Post-travail - Temps total"]
        ws.insert_cols(col_post + 1)
        ws.cell(1, col_post + 1).value = "Taux Post-travail"

    headers = {cell.value: cell.column for cell in ws[1]}

    if "Taux Post-travail" in headers:
        for r in range(2, ws.max_row + 1):
            post = to_seconds(ws.cell(r, headers.get("Post-travail - Temps total")).value)
            work = to_seconds(ws.cell(r, headers.get("Temps total Travail")).value)
            cell = ws.cell(r, headers["Taux Post-travail"])
            if post is None or work in (None, 0):
                cell.value = "-"
            else:
                cell.value = round(post / work, 4)
                cell.number_format = "0.00%"

    # add Moy Post-travail
    if "Taux Post-travail" in headers:
        col_taux = headers["Taux Post-travail"]
        ws.insert_cols(col_taux + 1)
        ws.cell(1, col_taux + 1).value = "Moy Post-travail"

    headers = {cell.value: cell.column for cell in ws[1]}

    if "Moy Post-travail" in headers and "Appels entrants" in headers:
        for r in range(2, ws.max_row + 1):
            post_sec = to_seconds(ws.cell(r, headers.get("Post-travail - Temps total")).value)
            appels_val = ws.cell(r, headers.get("Appels entrants")).value
            try:
                appels = float(appels_val) if appels_val not in ("-", None, "") else None
            except Exception:
                appels = None
            cell = ws.cell(r, headers["Moy Post-travail"])
            if post_sec is None or appels in (None, 0):
                cell.value = "-"
            else:
                cell.value = timedelta(seconds=int(round(post_sec / appels)))
                cell.number_format = "hh:mm:ss"

    # rule <10min presence
    if "Temps Total présence" in headers:
        col_presence = headers["Temps Total présence"]
        for r in range(2, ws.max_row + 1):
            sec = to_seconds(ws.cell(r, col_presence).value)
            if sec is not None and sec < 600:
                for c in range(col_presence, ws.max_column + 1):
                    ws.cell(r, c).value = "-"

    # styling
    header_fill = PatternFill("solid", "D9E1F2")
    alt_fill = PatternFill("solid", "F7F7F7")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for c in ws[1]:
        c.fill = header_fill
        c.font = Font(bold=True)

    for r in range(2, ws.max_row + 1):
        if r % 2 == 0:
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = alt_fill

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    for col in ws.columns:
        vals = [str(cell.value) for cell in col if cell.value not in (None, "")]
        ws.column_dimensions[get_column_letter(col[0].column)].width = max([len(v) for v in vals], default=10) + 2

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================
# CODE 5: Synthèse TL + Coaching + Counts + CF
# =========================
def code5_add_tl_sheet(excel_agent_bytes: bytes) -> bytes:
    from openpyxl.worksheet.cell_range import MultiCellRange

    wb = load_workbook(BytesIO(excel_agent_bytes))
    df = pd.read_excel(BytesIO(excel_agent_bytes), sheet_name="Flash Prod Agent", engine="openpyxl")

    if "Tls" not in df.columns:
        raise ValueError("Colonne 'Tls' introuvable dans Flash Prod Agent.")

    def to_sec_col(colname: str) -> pd.Series:
        if colname not in df.columns:
            return pd.Series([np.nan] * len(df))
        return pd.to_timedelta(df[colname].replace("-", np.nan), errors="coerce").dt.total_seconds()

    df["__presence_sec"] = to_sec_col("Temps Total présence")
    df["__work_sec"] = to_sec_col("Temps total Travail")
    df["__post_sec"] = to_sec_col("Post-travail - Temps total")
    df["__dmc_sec"] = to_sec_col("DMC")
    df["__dmt_sec"] = to_sec_col("DMT")

    df["__meeting_sec"] = to_sec_col("Meeting - Temps total")
    df["__training_sec"] = to_sec_col("Training - Temps total")
    df["__ojt_sec"] = to_sec_col("OJT - Temps total")
    df["__coaching_sec"] = df["__meeting_sec"].fillna(0) + df["__training_sec"].fillna(0) + df["__ojt_sec"].fillna(0)

    df["__prod"] = pd.to_numeric(df["Productivité"], errors="coerce") if "Productivité" in df.columns else np.nan

    present_mask = df["__presence_sec"].fillna(0) >= 600

    g = df.groupby("Tls", dropna=False)

    sum_presence = g["__presence_sec"].sum()
    sum_work = g["__work_sec"].sum()
    sum_post = g["__post_sec"].sum()

    sum_meeting = g["__meeting_sec"].sum()
    sum_training = g["__training_sec"].sum()
    sum_ojt = g["__ojt_sec"].sum()
    sum_coaching = g["__coaching_sec"].sum()

    agent_id_col = "Log Téléphonie1" if "Log Téléphonie1" in df.columns else ("Nom Agent" if "Nom Agent" in df.columns else None)

    if agent_id_col is None:
        agents_presents = df[present_mask].groupby("Tls").size()
        agents_lt13 = df[present_mask & df["__prod"].notna() & (df["__prod"] > 0) & (df["__prod"] < 13)].groupby("Tls").size()
    else:
        agents_presents = df[present_mask].groupby("Tls")[agent_id_col].nunique()
        agents_lt13 = df[present_mask & df["__prod"].notna() & (df["__prod"] > 0) & (df["__prod"] < 13)].groupby("Tls")[agent_id_col].nunique()

    idx = sum_presence.index

    df_tl = pd.DataFrame({
        "Tls": idx,
        "Agents présents": agents_presents.reindex(idx).fillna(0).astype(int),
        "Productivité": g["__prod"].mean(),
        "Agents < 13": agents_lt13.reindex(idx).fillna(0).astype(int),
        "Taux Post-travail": (sum_post / sum_work).replace([np.inf, -np.inf], np.nan),
        "Taux d'occupation": (sum_work / sum_presence).replace([np.inf, -np.inf], np.nan),
        "DMC": g["__dmc_sec"].mean(),
        "DMT": g["__dmt_sec"].mean(),
        "Heures Meeting": sum_meeting,
        "Heures Training": sum_training,
        "Heures OJT": sum_ojt,
        "Heures Coaching Total": sum_coaching,
        "% Coaching vs Connecté": (sum_coaching / sum_presence).replace([np.inf, -np.inf], np.nan),
    }).reset_index(drop=True)

    df_tl = df_tl[(df_tl["Agents présents"] > 0) | (df_tl["Productivité"].fillna(0) > 0)].copy()

    total_row = {
        "Tls": "TOTAL",
        "Agents présents": int(agents_presents.sum()),
        "Productivité": df_tl[df_tl["Tls"] != "TOTAL"]["Productivité"].mean(),
        "Agents < 13": int(agents_lt13.sum()),
        "Taux Post-travail": (sum_post.sum() / sum_work.sum()) if sum_work.sum() else np.nan,
        "Taux d'occupation": (sum_work.sum() / sum_presence.sum()) if sum_presence.sum() else np.nan,
        "DMC": df_tl[df_tl["Tls"] != "TOTAL"]["DMC"].mean(),
        "DMT": df_tl[df_tl["Tls"] != "TOTAL"]["DMT"].mean(),
        "Heures Meeting": sum_meeting.sum(),
        "Heures Training": sum_training.sum(),
        "Heures OJT": sum_ojt.sum(),
        "Heures Coaching Total": sum_coaching.sum(),
        "% Coaching vs Connecté": (sum_coaching.sum() / sum_presence.sum()) if sum_presence.sum() else np.nan,
    }
    df_tl = pd.concat([df_tl, pd.DataFrame([total_row])], ignore_index=True)

    for c in ["DMC", "DMT", "Heures Meeting", "Heures Training", "Heures OJT", "Heures Coaching Total"]:
        if c in df_tl.columns:
            df_tl[c] = pd.to_timedelta(df_tl[c], unit="s", errors="coerce")

    if "Flash Prod TL" in wb.sheetnames:
        del wb["Flash Prod TL"]
    ws = wb.create_sheet("Flash Prod TL")

    ws.append(df_tl.columns.tolist())
    for _, row in df_tl.iterrows():
        ws.append(row.tolist())

    header_fill = PatternFill("solid", "D9E1F2")
    total_fill = PatternFill("solid", "FFF2CC")
    green = PatternFill("solid", "C6EFCE")
    red = PatternFill("solid", "F4CCCC")
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = {c.value: c.column for c in ws[1]}

    formats = {
        "Agents présents": "0",
        "Productivité": "0.0",
        "Agents < 13": "0",
        "Taux Post-travail": "0.00%",
        "Taux d'occupation": "0.00%",
        "% Coaching vs Connecté": "0.00%",
        "DMC": "hh:mm:ss",
        "DMT": "hh:mm:ss",
        "Heures Meeting": "hh:mm:ss",
        "Heures Training": "hh:mm:ss",
        "Heures OJT": "hh:mm:ss",
        "Heures Coaching Total": "hh:mm:ss",
    }
    for k, fmt in formats.items():
        if k in headers:
            col = headers[k]
            for r in range(2, ws.max_row + 1):
                ws.cell(r, col).number_format = fmt

    for c in ws[1]:
        c.fill = header_fill
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for r in range(2, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center")

    last_row = ws.max_row
    if last_row >= 3:
        end_row = last_row - 1

        def add_cf(range_str: str, rule):
            ws.conditional_formatting.add(MultiCellRange(range_str), rule)

        if "Productivité" in headers:
            colL = get_column_letter(headers["Productivité"])
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2>13"], fill=green))
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2<=13"], fill=red))

        if "Taux Post-travail" in headers:
            colL = get_column_letter(headers["Taux Post-travail"])
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2<=0.08"], fill=green))
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2>0.08"], fill=red))

        for nm in ["DMC", "DMT"]:
            if nm in headers:
                colL = get_column_letter(headers[nm])
                add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2<=TIME(0,3,0)"], fill=green))
                add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2>TIME(0,3,0)"], fill=red))

        if "Taux d'occupation" in headers:
            colL = get_column_letter(headers["Taux d'occupation"])
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2>=0.7"], fill=green))
            add_cf(f"{colL}2:{colL}{end_row}", FormulaRule(formula=[f"{colL}2<0.7"], fill=red))

    for c in range(1, ws.max_column + 1):
        ws.cell(ws.max_row, c).fill = total_fill
        ws.cell(ws.max_row, c).font = Font(bold=True)

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18
    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================
# CODE 6: Email (MAJ coaching + counts)
# =========================
def code6_build_email(excel_final_bytes: bytes, projet_upper: str, date_txt: str, signature_name: str):
    df_tl = pd.read_excel(BytesIO(excel_final_bytes), sheet_name="Flash Prod TL", engine="openpyxl")

    total_row = df_tl[df_tl["Tls"] == "TOTAL"]
    total = total_row.iloc[0].to_dict() if not total_row.empty else {}

    kpi_prod = total.get("Productivité", np.nan)
    kpi_occ = total.get("Taux d'occupation", np.nan)
    kpi_post = total.get("Taux Post-travail", np.nan)
    kpi_coach = total.get("% Coaching vs Connecté", np.nan)

    dmc_sec = time_to_seconds(total.get("DMC", np.nan))
    dmt_sec = time_to_seconds(total.get("DMT", np.nan))

    subj_occ_txt = kpi_or_dash_pct(kpi_occ, 1)
    email_subject = f"{projet_upper} ==> Flash Prod AE de la journée du {date_txt} - Taux d'occupation {subj_occ_txt}"

    band_prod = kpi_or_dash_num(kpi_prod, 1)
    band_occ = kpi_or_dash_pct(kpi_occ, 1)
    band_post = kpi_or_dash_pct(kpi_post, 1)
    band_coach = kpi_or_dash_pct(kpi_coach, 1)
    band_dmc = sec_to_hms(dmc_sec)
    band_dmt = sec_to_hms(dmt_sec)

    df_tl_view = df_tl[df_tl["Tls"] != "TOTAL"].copy()

    def fmt_cell(c, v):
        if pd.isna(v):
            return "-"
        if c in ("Taux Post-travail", "Taux d'occupation", "% Coaching vs Connecté"):
            return f"{float(v)*100:.1f}%".replace(".", ",")
        if c in ("DMC", "DMT", "Heures Meeting", "Heures Training", "Heures OJT", "Heures Coaching Total"):
            return str(v) if str(v) != "NaT" else "-"
        if c == "Productivité":
            return f"{float(v):.1f}".replace(".", ",")
        if c in ("Agents présents", "Agents < 13"):
            try:
                return str(int(float(v)))
            except:
                return str(v)
        return str(v)

    table_html = """
    <table style="border-collapse:collapse;width:100%;font-size:13px;text-align:center;">
    <thead>
    <tr style="background:#203864;color:white;">
      <th>TL</th>
      <th>Agents présents</th>
      <th>Productivité</th>
      <th>Agents &lt; 13</th>
      <th>Taux Post-travail</th>
      <th>% Coaching</th>
      <th>DMC</th>
      <th>DMT</th>
      <th>Taux d’occupation</th>
    </tr>
    </thead><tbody>
    """
    for i, r in df_tl_view.iterrows():
        bg = "#F9FAFB" if i % 2 == 0 else "#FFFFFF"
        table_html += f"""
        <tr style="background:{bg};">
          <td>{r.get('Tls','-')}</td>
          <td>{fmt_cell('Agents présents', r.get('Agents présents', np.nan))}</td>
          <td>{fmt_cell('Productivité', r.get('Productivité', np.nan))}</td>
          <td>{fmt_cell('Agents < 13', r.get('Agents < 13', np.nan))}</td>
          <td>{fmt_cell('Taux Post-travail', r.get('Taux Post-travail', np.nan))}</td>
          <td>{fmt_cell('% Coaching vs Connecté', r.get('% Coaching vs Connecté', np.nan))}</td>
          <td>{fmt_cell('DMC', r.get('DMC', np.nan))}</td>
          <td>{fmt_cell('DMT', r.get('DMT', np.nan))}</td>
          <td>{fmt_cell("Taux d'occupation", r.get("Taux d'occupation", np.nan))}</td>
        </tr>
        """
    table_html += "</tbody></table>"

    signature_name = (signature_name or "").strip() or "MAHAMID Yassine"

    email_html = f"""
    <div style="background:#F3F6FB;padding:30px;font-family:Calibri,Arial;">
      <div style="max-width:950px;margin:auto;background:#FFFFFF;border-radius:14px;padding:26px;border:1px solid #E0E6ED;">

        <div style="border-left:5px solid #203864;padding-left:14px;margin-bottom:20px;">
          <div style="font-size:20px;font-weight:bold;color:#203864;">📊 Flash Production AE – {projet_upper}</div>
          <div style="font-size:13px;color:#6B7280;">Données du {date_txt}</div>
        </div>

        <table style="width:100%;margin-bottom:25px;border-radius:10px;background:#F9FAFB;text-align:center;font-size:13px;">
          <tr style="font-weight:bold;color:#203864;">
            <td>📈 Productivité<br><span style="font-size:18px;">{band_prod}</span></td>
            <td>⏱️ DMC<br><span style="font-size:18px;">{band_dmc}</span></td>
            <td>⏱️ DMT<br><span style="font-size:18px;">{band_dmt}</span></td>
            <td>🧩 Taux Post-travail<br><span style="font-size:18px;">{band_post}</span></td>
            <td>🎓 Coaching vs Connecté<br><span style="font-size:18px;">{band_coach}</span></td>
            <td style="background:#E8F5E9;border-radius:8px;color:#1E8449;">
              🎯 Taux d’occupation<br><span style="font-size:20px;font-weight:bold;">{band_occ}</span>
            </td>
          </tr>
        </table>

        <p>Bonjour,</p>
        <p>
          Vous trouverez ci-après les <b>réalisations KPI / Productivité par équipe AE</b> du <b>{date_txt}</b>,
          ainsi que le <b>détail par agent</b> en pièce jointe.
        </p>

        <p><b>➡️ Synthèse Productivité / équipe AE :</b></p>
        {table_html}

        <p style="margin-top:20px;">📎 <b>Pièce jointe :</b> Flash Prod AE – Détail Agents &amp; Synthèse TL</p>

        <p style="margin-top:20px;">
          Cordialement,<br><br>
          <b style="color:#1F4E78;">{signature_name}</b><br>
          FlashProd KPI Suite<br>
          <span style="color:#6B7280;font-size:12px;">Developed by MAHAMID Yassine</span>
        </p>

      </div>
    </div>
    """
    return email_subject, email_html


def copy_buttons(subject: str, html_body: str):
    component = f"""
    <div style="font-family:Calibri, Arial; margin-top:10px;">
      <div style="margin-bottom:6px; font-weight:bold;">📋 Copier vers Gmail :</div>

      <textarea id="sbj" rows="2" style="width:100%;">{subject}</textarea>
      <button style="margin-top:6px; padding:8px 14px; border:0; border-radius:6px; cursor:pointer; background:#28A745; color:white;"
              onclick="navigator.clipboard.writeText(document.getElementById('sbj').value)">
        ✅ Copier l'objet
      </button>

      <div style="height:10px;"></div>

      <div id="htmlBody" style="display:none;">{html_body}</div>
      <button style="padding:8px 14px; border:0; border-radius:6px; cursor:pointer; background:#0078D7; color:white;"
              onclick="navigator.clipboard.write([
                new ClipboardItem({
                  'text/html': new Blob([document.getElementById('htmlBody').innerHTML], {type:'text/html'})
                })
              ])">
        ✅ Copier le corps HTML (coller formaté dans Gmail)
      </button>
    </div>
    """
    st.components.v1.html(component, height=210, scrolling=False)


# =========================
# FLOP (optional)
# =========================
def build_flop_from_excel(excel_final_bytes: bytes, kpi_choice: str, flop_n: int, direction: str) -> pd.DataFrame:
    df_agents = pd.read_excel(BytesIO(excel_final_bytes), sheet_name="Flash Prod Agent", engine="openpyxl")
    base_cols = [c for c in ["Matricule RH", "Log Téléphonie1", "Nom Agent", "File", "Tls", "OPS"] if c in df_agents.columns]
    df = df_agents.copy()

    if "Productivité" in df.columns:
        df["__prod"] = pd.to_numeric(df["Productivité"], errors="coerce")
    if "Taux Post-travail" in df.columns:
        df["__tpost"] = df["Taux Post-travail"].apply(parse_percent_any)
    if "Taux d'occupation" in df.columns:
        df["__tocc"] = df["Taux d'occupation"].apply(parse_percent_any)
    if "DMC" in df.columns:
        df["__dmc_sec"] = df["DMC"].apply(parse_hms_any_to_seconds)
    if "DMT" in df.columns:
        df["__dmt_sec"] = df["DMT"].apply(parse_hms_any_to_seconds)
    if "Moy Post-travail" in df.columns:
        df["__mpost_sec"] = df["Moy Post-travail"].apply(parse_hms_any_to_seconds)

    mapping = {
        "Productivité": ("__prod", "number"),
        "Taux d'occupation": ("__tocc", "percent"),
        "Taux Post-travail": ("__tpost", "percent"),
        "DMC": ("__dmc_sec", "seconds"),
        "DMT": ("__dmt_sec", "seconds"),
        "Moy Post-travail": ("__mpost_sec", "seconds"),
    }
    score_col, kind = mapping[kpi_choice]
    if score_col not in df.columns:
        raise ValueError(f"KPI '{kpi_choice}' introuvable dans Flash Prod Agent.")

    if "Temps Total présence" in df.columns:
        pres_sec = df["Temps Total présence"].apply(parse_hms_any_to_seconds)
        df = df[(pres_sec.isna()) | (pres_sec >= 600)].copy()

    df = df[df[score_col].notna()].copy()

    # auto: productivité et occupancy => worst lowest, post/dmc/dmt => worst highest
    if direction == "Worst (auto)":
        direction = "Worst (lowest)" if kpi_choice in ["Productivité", "Taux d'occupation"] else "Worst (highest)"

    ascending = True if direction == "Worst (lowest)" else False
    df_sorted = df.sort_values(score_col, ascending=ascending).head(int(flop_n)).copy()

    def fmt_value(v):
        if pd.isna(v):
            return "-"
        if kind == "percent":
            return f"{float(v) * 100:.1f}%".replace(".", ",")
        if kind == "seconds":
            return sec_to_hms(float(v))
        return str(round(float(v), 2)).replace(".", ",")

    df_sorted["KPI"] = kpi_choice
    df_sorted["Valeur KPI"] = df_sorted[score_col].apply(fmt_value)

    out_cols = base_cols + ["KPI", "Valeur KPI"]
    return df_sorted[out_cols]


def flop_to_excel_bytes(df_flop: pd.DataFrame, sheet_name: str = "FLOP") -> bytes:
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df_flop.to_excel(writer, index=False, sheet_name=sheet_name)
    buf.seek(0)
    return buf.getvalue()


# =========================
# UI
# =========================
st.title("📊 FlashProd KPI Suite")
st.caption("Pipeline BRUT + COMPO → KPI Agents + Synthèse TL (Coaching/Counts) → Excel final + Email HTML + FLOP (optionnel)")

with st.sidebar:
    st.header("Configuration")
    signature_name = st.text_input("Nom signature email", value="MAHAMID Yassine")
    projet = st.text_input("Nom du projet (ex: CNSS)", value="CNSS")
    date_input = st.text_input("Date Flash Prod (JJ/MM/AAAA)", value=datetime.today().strftime("%d/%m/%Y"))

    st.markdown("---")
    st.subheader("Uploads")
    raw_file = st.file_uploader("📁 Fichier BRUT (.xls / .xlsx)", type=["xls", "xlsx"])
    compo_file = st.file_uploader("📁 Fichier COMPO (.xls / .xlsx)", type=["xls", "xlsx"])

    st.markdown("---")
    st.subheader("FLOP (optionnel)")
    enable_flop = st.checkbox("Générer un fichier FLOP (worst performers)", value=False)
    kpi_choice = st.selectbox(
        "KPI pour FLOP",
        ["Productivité", "Taux d'occupation", "Taux Post-travail", "DMC", "DMT", "Moy Post-travail"],
        index=0,
        disabled=not enable_flop
    )
    flop_n = st.number_input("Nombre d'agents FLOP", min_value=5, max_value=200, value=20, step=5, disabled=not enable_flop)
    direction_mode = st.radio(
        "Tri FLOP",
        ["Worst (auto)", "Worst (lowest)", "Worst (highest)"],
        index=0,
        disabled=not enable_flop
    )

    st.markdown("---")
    run = st.button("🚀 Lancer le pipeline", type="primary")

if not run:
    st.info("Uploade BRUT + COMPO puis clique **Lancer le pipeline**.")
    st.markdown("<div style='text-align:center;color:#6c757d;font-size:12px;margin-top:16px;'>Developed by <b>MAHAMID Yassine</b></div>", unsafe_allow_html=True)
    st.stop()

if not raw_file or not compo_file:
    st.error("BRUT + COMPO sont obligatoires.")
    st.stop()

try:
    date_obj = datetime.strptime(date_input.strip(), "%d/%m/%Y")
except Exception:
    st.error("Format date invalide. Utilise JJ/MM/AAAA (ex: 09/02/2026).")
    st.stop()

projet_upper = projet.strip().upper()
date_txt = date_obj.strftime("%d/%m/%Y")
date_flash = date_obj.strftime("%d_%m_%Y")

status_box = st.empty()
logs = []

def log_ok(msg):
    logs.append(msg)
    status_box.success("\n".join(logs))

try:
    df_raw = read_excel_any(raw_file)
    df_compo = read_excel_any(compo_file)

    df_clean = code1_clean_raw(df_raw)
    log_ok("✅ CODE 1 OK → Nettoyage BRUT terminé")

    extracted_logs = df_clean["Log Téléphonie1"].notna().sum() if "Log Téléphonie1" in df_clean.columns else 0
    st.info(f"🔎 BRUT: logs extraits = {extracted_logs} / lignes = {len(df_clean)}")

    df_rapport = code2_build_agent_kpis(df_clean)
    log_ok("✅ CODE 2 OK → KPI Agent / Pivot terminé")

    df_final3, merge_stats = code3_merge_compo(df_compo, df_rapport)
    log_ok("✅ CODE 3 OK → Merge COMPO + Rapport terminé")

    st.info(
        f"🔎 MERGE: matched = {merge_stats['matched_rows']} | unmatched = {merge_stats['unmatched_rows']} "
        f"(COMPO non-null logs = {merge_stats['compo_rows_non_null_log']}, rapport rows = {merge_stats['rapport_rows']})"
    )

    agent_excel_bytes = code4_build_flash_agent_excel(df_final3)
    log_ok("✅ CODE 4 OK → Flash Prod Agent (Excel) généré")

    excel_final_bytes = code5_add_tl_sheet(agent_excel_bytes)
    excel_final_name = f"Flash_Prod_{projet_upper}_{date_flash}.xlsx"
    log_ok(f"✅ CODE 5 OK → Excel final prêt ({excel_final_name})")

    email_subject, email_html = code6_build_email(excel_final_bytes, projet_upper, date_txt, signature_name)
    log_ok("✅ CODE 6 OK → Email (Objet + Corps HTML) généré")

except Exception as e:
    st.error(f"Erreur pipeline: {e}")
    st.stop()

df_flop = None
flop_excel_bytes = None
flop_filename = None

if enable_flop:
    try:
        df_flop = build_flop_from_excel(excel_final_bytes, kpi_choice, int(flop_n), direction_mode)
        flop_excel_bytes = flop_to_excel_bytes(df_flop, sheet_name="FLOP")
        flop_filename = f"FLOP_{kpi_choice.replace(' ', '_')}_{projet_upper}_{date_flash}.xlsx"
    except Exception as e:
        st.warning(f"FLOP non généré: {e}")

col1, col2 = st.columns([1.15, 0.85], gap="large")

with col1:
    st.subheader("✉️ Email")
    st.text_input("Objet", value=email_subject)
    st.components.v1.html(email_html, height=560, scrolling=True)
    copy_buttons(email_subject, email_html)

with col2:
    st.subheader("⬇️ Fichiers")
    st.download_button(
        label="Télécharger l'Excel final",
        data=excel_final_bytes,
        file_name=excel_final_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    if enable_flop and flop_excel_bytes is not None:
        st.download_button(
            label="Télécharger le fichier FLOP",
            data=flop_excel_bytes,
            file_name=flop_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    st.subheader("👀 Aperçu – Flash Prod TL (Top 50)")
    try:
        df_tl_preview = pd.read_excel(BytesIO(excel_final_bytes), sheet_name="Flash Prod TL", engine="openpyxl")
        st.dataframe(df_tl_preview.head(50), use_container_width=True, height=300)
    except Exception as e:
        st.warning(f"Aperçu TL non disponible: {e}")

    if enable_flop and df_flop is not None:
        st.subheader(f"👎 Aperçu FLOP – {kpi_choice} (Top {int(flop_n)})")
        st.dataframe(df_flop, use_container_width=True, height=300)

st.markdown("<div style='text-align:center;color:#6c757d;font-size:12px;margin-top:16px;'>Developed by <b>MAHAMID Yassine</b></div>", unsafe_allow_html=True)
